import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:

    @staticmethod
    def exporter_grand_livre(data, nom_societe, code_exercice):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Grand Livre'

        # En-tête
        ws['A1'] = f'GRAND LIVRE - {nom_societe} - Exercice {code_exercice}'
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        for compte in data.get('comptes', []):
            # En-tête compte
            ws.cell(row=row, column=1, value=f"{compte['numero']} - {compte['intitule']}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # En-tête colonnes
            headers = ['Date', 'Journal', 'Pièce', 'Libellé', 'Débit', 'Crédit', 'Lettrage', 'Solde']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            row += 1

            for l in compte.get('lignes', []):
                ws.cell(row=row, column=1, value=str(l['date']))
                ws.cell(row=row, column=2, value=l['journal'])
                ws.cell(row=row, column=3, value=l['piece'])
                ws.cell(row=row, column=4, value=l['libelle'])
                ws.cell(row=row, column=5, value=float(l['debit']))
                ws.cell(row=row, column=6, value=float(l['credit']))
                ws.cell(row=row, column=7, value=l['lettrage'])
                ws.cell(row=row, column=8, value=float(l['solde']))
                row += 1

            # Total compte
            ws.cell(row=row, column=4, value='TOTAL')
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=5, value=float(compte['total_debit']))
            ws.cell(row=row, column=5).font = Font(bold=True)
            ws.cell(row=row, column=6, value=float(compte['total_credit']))
            ws.cell(row=row, column=6).font = Font(bold=True)
            row += 2

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def exporter_balance(data, nom_societe, code_exercice):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Balance'

        ws['A1'] = f'BALANCE - {nom_societe} - Exercice {code_exercice}'
        ws['A1'].font = Font(bold=True, size=14)

        headers = ['Compte', 'Intitulé', 'Total Débit', 'Total Crédit', 'Solde Débiteur', 'Solde Créditeur']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        row = 4
        for ligne in data.get('lignes', []):
            ws.cell(row=row, column=1, value=ligne['numero'])
            ws.cell(row=row, column=2, value=ligne['intitule'])
            ws.cell(row=row, column=3, value=float(ligne['total_debit']))
            ws.cell(row=row, column=4, value=float(ligne['total_credit']))
            ws.cell(row=row, column=5, value=float(ligne['solde_debiteur']))
            ws.cell(row=row, column=6, value=float(ligne['solde_crediteur']))
            row += 1

        # Totaux
        ws.cell(row=row, column=2, value='TOTAL')
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data.get('total_debit', 0)))
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(data.get('total_credit', 0)))
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(data.get('total_solde_debiteur', 0)))
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(data.get('total_solde_crediteur', 0)))
        ws.cell(row=row, column=6).font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
